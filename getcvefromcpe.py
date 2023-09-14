import openpyxl
import nvdlib
import re
import sys
import json
import requests

def get_cpe_from_buildrt(jfile):

    cpe_strings = [] 
    
    with open(jfile, "r") as file:
       data = json.load(file)

    # Extract details for each package
    for package_name, package_details in data["packages"].items():
        current_version = package_details.get("current_version", "N/A")
        license_string = package_details.get("license", "N/A")
        cpe_id = package_details.get("cpeid", "N/A")
        print(f"Package Name: {package_name}")
        print(f"license_string: {license_string}")
        print(f"Current Version: {current_version}")
        #print(f"CPE ID: {cpe_id}")
        #print("-" * 40)
        if not cpe_id:
            continue
        else: 
            cpe_strings.append(str(cpe_id))

    return cpe_strings



def get_epss_scores(cve):
    epss_scores = {}
    
    try:
        # Get the EPSS Score from the FIRST EPSS API
        response = requests.get('https://api.first.org/data/v1/epss/?cve=' + cve.strip())
        
        if response.status_code == 200:
            data = response.json()
            
            for item in data['data']:
                cve = item["cve"]
                epss = item["epss"]
                percentile = item["percentile"]
                
                epss_scores[cve] = {
                    'epss': epss,
                    'percentile': percentile
                }
                
                #print(f"CVE: {cve}, EPSS Score: {epss}, Percentile: {percentile} ")
        else:
            epss_scores[cve] = 'Error: Status code ' + str(response.status_code)
    except Exception as e:
        epss_scores[cve] = 'Error: ' + str(e)
    
    return epss_scores


def validate_input(input_str):
    # Regular expression to check for spaces or special characters
    pattern = re.compile(r'[\s!@#$%^&*()=_+[\]{}|;:",.<>?/\\]')

    # Check if the input contains spaces or special characters
    if pattern.search(input_str):
        # If spaces are encountered, ignore the rest of the input
        input_str = input_str.split()[0]

    return input_str


def validate_version_string(s):
    
    s = re.sub(r'\s+', '', s)  # Remove all whitespace characters from the string

    # Regular expression patterns for each format
    patterns = [
        r'^\d+\.\d+\.\d+[a-zA-Z]$',  # 1.1.1a
        r'^\d+\.\d+[a-zA-Z]$',      # 2.3a
        r'^\d+\.\d+$',              # 2.3
        r'^\d+$',                   # 256
        r'^\d+\.\d+\.\d+\.\d+$',    # 0.0.0.0
        r'^\d+\.\d+-RC\d+$',        # 2020.1-RC1
        r'^\d+\.\d+\.\d+$',         # 5.4.234
        r'^[a-fA-F0-9]{40}$',        # 40-character hexadecimal 
        r'^\d+\.\d+\.\d+-\d{14}-[a-fA-F0-9]+$' # 0.0.0-20200622213623-75b288015ac9
    ]
    
    
    return any(re.match(pattern, s) for pattern in patterns)



def process_nums(input_str):
    valid_input = []
    for char in input_str:
        if validate_version_string(input_str):
            valid_input.append(char)
        else:
            break
    return ''.join(valid_input)



def generate_cpe_string(part_name, package_name, vendor, package_version, update_version):
    cpe_version = "2.3"  # CPE version
    part = part_name # Part is typically 'a' for applications
    product = package_name
    version = package_version
    update = update_version
    edition = "*"
    language = "*"
    sw_edition = "*"
    target_sw = "*"
    target_hw = "*"
    other = "*"

    cpe_string = f"cpe:{cpe_version}:{part}:{vendor}:{product}:{version}:{update}:{edition}:{language}:{sw_edition}:{target_sw}:{target_hw}:{other}"
    #cpe_string = cpe_string.replace("::", ":*:")

    pattern = re.compile(r'\s+')
    cpe_string = re.sub(pattern, '', cpe_string)
    
    return cpe_string

def read_packages_from_excel(filename, sheet):
    package_data = []
    
    part_column = None
    package_column = None
    vendor_column = None
    version_column = None
    update_column = None
    
    try:
        workbook = openpyxl.load_workbook(filename)
      
        sheet = workbook[sheet]
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for index, cell_value in enumerate(row):
                if cell_value == 'Package':
                    package_column = index + 1
                elif cell_value == 'Part':
                    part_column = index + 1
                elif cell_value == 'Vendor':
                    vendor_column = index + 1
                elif cell_value == 'Version':
                    version_column = index + 1
                elif cell_value == 'Update':
                    update_column = index + 1
                        
        if part_column is None or package_column is None or vendor_column is None or version_column is None:
            raise ValueError("Column headers 'Part', 'Package', 'Vendor', and 'Version' not found.")
            
        for row in sheet.iter_rows(min_row=2, values_only=True):
          
            part_name = row[part_column - 1]
            if not part_name:
                part_name = "a"
            part_name = validate_input(str(part_name))
            
            package_name = row[package_column - 1]
            package_name = validate_input(str(package_name))
            vendor = row[vendor_column - 1]
            if not vendor:
                vendor = package_name
            vendor = validate_input(str(vendor))
            package_version = row[version_column - 1]
            if not package_version:
                print(f"version missing for {vendor} ")
                continue

            package_version = process_nums(str(package_version))
            update_version = row[update_column - 1]
            if not update_version:
                update_version = "*"
            else: 
                update_version = validate_input(str(update_version))    
            cpe_string = generate_cpe_string(part_name, package_name, vendor, package_version, update_version)
            package_data.append(cpe_string)
                    
    except Exception as e:
        print(f"An error occurred: {e}")
        
    return package_data
 

if __name__ == "__main__":


    if len(sys.argv) == 1:

        #get package information fron buildroot return cves and licence info
        cpe_strings = get_cpe_from_buildrt("pkg-stats.json")
        for cpe_id in cpe_strings:
            print(cpe_id)
            r = nvdlib.searchCVE(cpeName = cpe_id)
            for eachCVE in r:
                    epss_scores = get_epss_scores(eachCVE.id)
                    print(eachCVE.id, eachCVE.score, epss_scores[eachCVE.id], eachCVE.url)        
"""
        #get information from spreadsheet return cves
        excel_filename = "Book1.xlsx"

        #modify the list as appropriate
        #sheet_name = ['product1', 'product2', 'product3','product4','product5']
        sheet_name = ['product9']

        cve_list = []

   
        for sheet in sheet_name:
            print(sheet)
            cpe_strings = read_packages_from_excel(excel_filename, sheet)
            for cpe_string in cpe_strings:
                print(f"CPE String: {cpe_string}")
                r = nvdlib.searchCVE(cpeName = cpe_string)
                for eachCVE in r:
                    epss_scores = get_epss_scores(eachCVE.id)
                    print(eachCVE.id, eachCVE.score, epss_scores[eachCVE.id], eachCVE.url)
 
    else:
        #get test string from commandline - feed in the specific cpe
        arguments = sys.argv[1:]
        for arg in arguments:
            print ("command args:", arg) 
            r = nvdlib.searchCVE(cpeName = arg)
            for eachCVE in r:
                epss_scores = get_epss_scores(eachCVE.id)
                print(eachCVE.id, eachCVE.score, epss_scores[eachCVE.id], eachCVE.url)
  
"""